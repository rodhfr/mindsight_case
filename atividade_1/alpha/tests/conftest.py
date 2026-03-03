"""
conftest.py — fixtures compartilhadas entre todos os módulos de teste.
"""

import io

import numpy as np
import openpyxl
import pandas as pd
import pytest


# ── DataFrames sintéticos ─────────────────────────────────────────────────────

@pytest.fixture
def df_perf_simples():
    """DataFrame mínimo com 4 funcionários e 2 semestres de performance."""
    return pd.DataFrame({
        "Nome Completo":   ["Ana", "Bruno", "Carla", "Diego"],
        "Área":            ["Comercial", "Operações", "Comercial", "TI"],
        "Potencial Bruto": [80, 40, 60, np.nan],
        "perf_1_2018":     [3.0, 1.0, 2.0, np.nan],
        "perf_2_2018":     [3.0, 2.0, 2.0, 1.0],
        "Raciocínio":      [70, 50, 60, 80],
        "Social":          [60, 55, 75, 65],
        "Motivacional":    [80, 45, 70, 55],
        "Cultura pontuação": [55, 30, 42, 68],
        "Cultura classificação": ["Medio-Alto", "Baixo", "Medio", "Alto"],
        "atributo-Comunicação": [70, 50, 65, 80],
        "atributo-Liderança":   [80, 40, 60, 75],
    })


@pytest.fixture
def df_sem_potencial(df_perf_simples):
    """Mesmo DataFrame sem a coluna Potencial Bruto."""
    return df_perf_simples.drop(columns=["Potencial Bruto"])


@pytest.fixture
def df_grande():
    """DataFrame maior (50 linhas) para testes que exigem n mínimo."""
    rng = np.random.default_rng(42)
    n = 50
    return pd.DataFrame({
        "perf_1_2018": rng.uniform(1, 3, n),
        "perf_2_2018": rng.uniform(1, 3, n),
        "Potencial Bruto": rng.uniform(0, 100, n),
        "Raciocínio":  rng.uniform(0, 100, n),
        "Social":      rng.uniform(0, 100, n),
        "atributo-Comunicação": rng.uniform(0, 100, n),
        "Cultura pontuação":    rng.uniform(0, 100, n),
    })


# ── Arquivos Excel sintéticos ─────────────────────────────────────────────────

def _criar_excel_valido() -> io.BytesIO:
    """Cria um workbook Excel válido com as 3 abas e colunas obrigatórias."""
    wb = openpyxl.Workbook()

    # Tablib Dataset
    ws_main = wb.active
    ws_main.title = "Tablib Dataset"
    colunas_main = [
        "Nome", "Sobrenome", "E-mail", "CPF",
        "Raciocínio", "Social", "Motivacional",
        "Cultura pontuação", "Cultura classificação",
        "Potencial Bruto", "Match",
        "perfil-Comunicação", "atributo-Comunicação",
        "URL-teste", "Início-teste",
    ]
    ws_main.append(colunas_main)

    # performance
    ws_perf = wb.create_sheet("performance")
    ws_perf.append(["CPF", "Performance 1-2018", "Performance 2-2018"])

    # área
    ws_area = wb.create_sheet("área")
    ws_area.append(["CPF", "Área"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _criar_excel_sem_abas() -> io.BytesIO:
    """Workbook com apenas a aba padrão (Sheet) — faltam as 3 abas esperadas."""
    wb = openpyxl.Workbook()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _criar_excel_sem_colunas() -> io.BytesIO:
    """Workbook com as 3 abas mas sem as colunas obrigatórias."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tablib Dataset"
    ws.append(["coluna_irrelevante"])
    wb.create_sheet("performance").append(["coluna_irrelevante"])
    wb.create_sheet("área").append(["coluna_irrelevante"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@pytest.fixture
def excel_valido():
    return _criar_excel_valido()


@pytest.fixture
def excel_sem_abas():
    return _criar_excel_sem_abas()


@pytest.fixture
def excel_sem_colunas():
    return _criar_excel_sem_colunas()


@pytest.fixture
def arquivo_invalido():
    """BytesIO com conteúdo que não é Excel."""
    return io.BytesIO(b"isso nao e um excel")
